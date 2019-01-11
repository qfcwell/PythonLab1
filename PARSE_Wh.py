# -*- coding: utf-8 -*-
import pymssql
from openpyxl import load_workbook

'''
def run():
    with pymssql.connect(host='10.1.246.1', user="sa", password="sasasa",database="CAPOL_Project") as conn:
        cur=conn.cursor()
        UserId=GetUserId(cur,'libina')
        print(UserId)
        AlterUserName(cur,UserId,'李斌2')
        AlterLoginName(cur,UserId,'libin2')
        conn.commit()
'''
conns={
    'shenzhen':('10.1.246.1', "sa", "sasasa","CAPOL_Core"),
    'shenzhen_old':('10.1.1.118', "sa", "hygj!@#456","TJA_Core"),
    'guangzhou':('10.2.1.114', "sa", "hygj!@#456","CAPOL_Core"),
    'hainan':('10.14.2.10',"sa", "hygj!@#456","CAPOL_Core"),
    'changsha':('10.3.1.3',"sa", "abcd.1234","CAPOL_Core"),
    'changsha_old':('10.3.1.3',"sa", "abcd.1234","TJA_Core"),
    'shanghai':('10.6.1.15',"sa", "capol!@#456","CAPOL_Core"),
    'shanghai_old':('10.6.1.15',"sa", "capol!@#456","TJA_Core"),
    }

servers=['shenzhen',
    'shenzhen_old',
    'guangzhou',
    'hainan',
    'changsha',
    'changsha_old',
    'shanghai',
    'shanghai_old']

file='D:\\3.xlsx'

def run():
    res=Parse(file)
    with pymssql.connect(host='10.1.42.131', user="sa", password="qfc23834358Q",database="test2") as conn:
        cur=conn.cursor()
        for item in res:
            cur.execute(u"INSERT INTO WH_import(prjno,prjname,prjstatus,total_budget,wh_fillin,total_rate,chushe_rate,shigongtu_rate) values(%s,%s,%s,%s,%s,%s,%s,%s)",item)
        conn.commit()

def Parse(file):
    res=[]
    workbook=load_workbook(file)
    sheet_name=workbook.sheetnames[1]
    worksheet=workbook[sheet_name]
    row=3
    while 1:
        prjno,prjname,prjstatus=worksheet.cell(row, 2).value,worksheet.cell(row,3).value,worksheet.cell(row, 5).value
        total_budget,wh_fillin=worksheet.cell(row, 6).value,worksheet.cell(row, 7).value
        total_rate,chushe_rate=shigongtu_rateworksheet.cell(row, 8).value,worksheet.cell(row, 9).value,worksheet.cell(row, 10).value
        row+=1
        if not prjno:
            break
        else:
            res.append((prjno,prjname,prjstatus,total_budget,wh_fillin,total_rate,chushe_rate,shigongtu_rate))
            print((prjno,prjname,prjstatus,total_budget,wh_fillin,total_rate,chushe_rate,shigongtu_rate))
    return res

if __name__=="__main__":
    run()
