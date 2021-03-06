﻿# -*- coding: utf-8 -*-
import pymssql
from openpyxl import load_workbook
conns={
    'shenzhen':('10.1.246.1', "sa", "sasasa","CAPOL_Core"),
    'shenzhen_old':('10.1.1.118', "sa", "hygj!@#456","TJA_Core"),
    'guangzhou':('10.2.1.114', "sa", "hygj!@#456","CAPOL_Core"),
    'hainan':('10.14.2.10',"sa", "hygj!@#456","CAPOL_Core"),
    'changsha':('10.3.1.3',"sa", "abcd.1234","CAPOL_Core"),
    'changsha_old':('10.3.1.3',"sa", "abcd.1234","TJA_Core"),
    'shanghai':('10.6.1.15',"sa", "capol!@#456","CAPOL_Core"),
    'shanghai_old':('10.6.1.15',"sa", "capol!@#456","TJA_Core"),
    }

servers=['shenzhen',
    'shenzhen_old',
    'guangzhou',
    'hainan',
    'changsha',
    'changsha_old',
    'shanghai',
    'shanghai_old']

def run():
    servers=[
    'changsha',
    ]
    for server in servers:
        with ACS_UserName_Tool(server=server) as app:
            #UserId=app.GetUserId(LoginName='tangzhi')
            #app.UpdateUserName(UserId=UserId,New_LoginName='tangzhi1',New_UserName='唐智1')
            #res=app.CheckUserName()

class ACS_UserName_Tool():
    def __init__(self,server):
        if server in conns:
            (host,user,password,database)=conns[server]
            self.conn=pymssql.connect(host=host, user=user, password=password,database=database)
        else:
            print("server name not exsist")
            self.conn=0

    def __enter__(self):
        return self

    def __exit__(self, type, value, trace):
        self.close()

    def close(self):
        self.conn.close()

    def CheckUserName(self):
        cur=self.conn.cursor()
        cur.execute("SELECT distinct su.id,su.UserName,su.LoginName from capol_project.dbo.designfile df join capol_core.dbo.sysuser su on df.LockerId=su.Id where df.LockerName<>su.UserName and su.LoginName is not null")
        res=cur.fetchall()
        if res:
            for (UserId,New_UserName,New_LoginName) in res:
                print((UserId,New_UserName,New_LoginName))
                self.UpdateUserName(UserId,New_LoginName,New_UserName)

    def GetUserId(self,LoginName):
        cur=self.conn.cursor()
        cur.execute(u"SELECT ID,UserName FROM SysUser where LoginName=%s",(LoginName))
        res=cur.fetchone()
        if res:
            (UserId,UserName)=res
            return UserId
        else:
            return 0

    def AlterUserName(self,UserId,NewName):#改中文姓名
        cur=self.conn.cursor()
        cur.execute(u"SELECT UserName FROM CAPOL_Core.dbo.SysUser where ID=%s",(UserId))
        (OldName,)=cur.fetchone()
        cur.execute(u"UPDATE SysUser set UserName=%s where ID=%s",(NewName,UserId))
        cur.execute(u"UPDATE CAPOL_Project.dbo.DesignFile set CreatorName=%s where CreatorId=%s",(NewName,UserId))#刷设计文件
        cur.execute(u"UPDATE CAPOL_Project.dbo.DesignFile set LastBackuperName=%s where LastBackuperId=%s",(NewName,UserId))#刷设计文件
        cur.execute(u"UPDATE CAPOL_Project.dbo.DesignFile set LockerName=%s where LockerId=%s",(NewName,UserId))#刷设计文件
        cur.execute(u"UPDATE CAPOL_Project.dbo.TakeoutFile set TakeoutUserName=%s where TakeoutUserId=%s",(NewName,UserId))#刷提资文件
        cur.execute(u"UPDATE CAPOL_Project.dbo.ProductFile set CreatorName=%s where CreatorId=%s",(NewName,UserId))#刷归档文件
        cur.execute(u"SELECT Id,ExecutorId,ExecutorName from CAPOL_Project.dbo.PrjRole WHERE ExecutorId is not null and ExecutorId!=''")
        for (RoleId,ExecutorId,ExecutorName) in cur.fetchall():
            if UserId in ExecutorId.split(','):
                lst=[]
                for uid in ExecutorId.split(','):
                    cur.execute(u"SELECT UserName FROM SysUser WHERE Id=%s",(uid))
                    (uname,)=cur.fetchone()
                    lst.append(uname)
                NewExecutorName=','.join(lst)
                cur.execute(u"UPDATE CAPOL_Project.dbo.PrjRole SET ExecutorName=%s where Id=%s",(NewExecutorName,RoleId))
        self.conn.commit()

    def AlterLoginName(self,UserId,NewName):
        cur=self.conn.cursor()
        cur.execute(u"UPDATE SysUser set LoginName=%s where ID=%s",(NewName,UserId))#修改用户名
        self.conn.commit()

    def UpdateUserName(self,UserId,New_LoginName,New_UserName):
        cur=self.conn.cursor()
        if UserId:
            print(UserId)
            self.AlterUserName(UserId,New_UserName)
            self.AlterLoginName(UserId,New_LoginName)
            self.conn.commit()
        else:
            print('无此用户')

file='D:\\桌面\\平台\\人员名单.xlsx'
'''
def run():
    for server in servers:
        (host,user,password,database)=conns[server]
        with pymssql.connect(host=host, user=user, password=password,database=database) as conn:
            cur=conn.cursor()
            i=0
            for (OldName,NewName) in GetUserList(file,5,3):
                UserId=GetUserId(cur,OldName)
                if UserId:
                    AlterLoginName(cur,UserId,NewName)
                    print((server,UserId,OldName,NewName))
                    i+=1
            print(server+':'+str(i))
            #conn.commit()

'''





def TestConnection():
    for server in servers:
        (host,user,password,database)=conns[server]
        with pymssql.connect(host=host, user=user, password=password,database=database) as conn:
            cur=conn.cursor()
            cur.execute(u"SELECT ID FROM SYSUSER")
            res=cur.fetchone()
            print(server)
            print(res)

def GetUserList(file,col1,col2):
    res=[]
    workbook=load_workbook(file)
    sheet_name=workbook.sheetnames[0]
    worksheet=workbook[sheet_name]
    row=2
    while 1:
        OldName,NewName= worksheet.cell(row, col1).value,worksheet.cell(row, col2).value
        row+=1
        if not OldName:
            break
        else:
            res.append((OldName,NewName))
    return res

if __name__=="__main__":
    run()
