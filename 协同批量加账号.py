# -*- coding: utf-8 -*- 运行python
import pymssql
from openpyxl import load_workbook
import os

def get_group_id(cur,group_name):
    cur.execute(u"SELECT Id FROM SysGroup WHERE GroupName=%s",(group_name))
    (group_id,)=cur.fetchone()
    return group_id

def get_role_id(cur,role_name):
    cur.execute(u"SELECT Id FROM SysRole WHERE RoleName=%s",(role_name))
    (role_id,)=cur.fetchone()
    return role_id

def get_user_list(file,col1,col2):
    res=[]
    workbook=load_workbook(file)
    sheet_name=workbook.sheetnames[0]
    worksheet=workbook[sheet_name]
    row=2
    while 1:
        user_name,login_name= worksheet.cell(row, col1).value,worksheet.cell(row, col2).value
        row+=1
        if not user_name:
            break
        else:
            res.append((user_name,login_name))
    return res

def generate_new_user_id(cur):
    cur.execute(u"exec  [dbo].[GetRecordId] 'Admin.SysUser'")
    cur.execute(u"SELECT Prefix,SequNo FROM SysGlobalId WHERE IdKey ='Admin.SysUser'")
    (Prefix,SequNo)=cur.fetchone()
    return Prefix+SequNo

def add_user(cur,login_name,user_name):
    user_id=generate_new_user_id(cur)
    cur.execute(u"INSERT INTO SysUser(Id,LoginName,UserName) values(%s,%s,%s)",(user_id,login_name,user_name))
    return user_id

def add_user_group(cur,user_id,group_id):
    cur.execute(u"SELECT COUNT(*) FROM SysGroupUser where UserId=%s",(user_id))
    (res,)=cur.fetchone()
    if res:
        print("UserId已存在："+user_id)
    else:
        cur.execute(u"INSERT INTO SysGroupUser(UserId,GroupId,IsDefault) values(%s,%s,'T')",(user_id,group_id))
    return 1

def add_user_role(cur,user_id,role_id):
    cur.execute(u"SELECT COUNT(*) FROM SysUserRole where UserId=%s and RoleId=%s",(user_id,role_id))
    (res,)=cur.fetchone()
    if res:
        print("角色已存在："+user_id)
    else:
        cur.execute(u"INSERT INTO SysUserRole(UserId,RoleId) values(%s,%s)",(user_id,role_id))
    return 1

def user_duplication_check(cur,login_name,user_name):
    cur.execute(u"SELECT count(*) from SysUser where UserName=%s or LoginName=%s",(user_name,login_name))
    (res,)=cur.fetchone()
    if res:
        print("登陆名或用户名重复："+login_name+","+user_name)
        return 0
    else:
        return 1

def run():
    #with pymssql.connect(host="10.1.246.1", user="sa", password="sasasa",database="CAPOL_Core") as conn:  #深圳新协同
    #with pymssql.connect(host="10.1.1.118", user="sa", password="hygj!@#456",database="TJA_Core") as conn: #深圳老协同
        #file=r"D:\桌面\校招生140人协同.xlsx"
        #lst=get_user_list(file,1,2)
    #with pymssql.connect(host="10.2.1.114", user="sa", password="hygj!@#456",database="CAPOL_Core") as conn: #广州新协同
        #file=r"D:\桌面\广州公司协同账号开通20180709.xlsx"
        #lst=get_user_list(file,4,5)
    #with pymssql.connect(host="10.3.1.3", user="sa", password="abcd.1234",database="CAPOL_Core") as conn: #长沙新协同
    #    file=r"D:\桌面\长沙公司新建账户.xlsx"
    #    lst=get_user_list(file,2,1)
    #with pymssql.connect(host="10.3.1.5", user="sa", password="fengfeiying@1128",database="TJA_Core") as conn: #长沙老协同
    #    file=r"D:\桌面\长沙公司新建账户.xlsx"
    #    lst=get_user_list(file,2,1)
        #lst=[('黄伟5','huangwei5')]
        #cur=conn.cursor()
        group_id=get_group_id(cur,'建筑施工图室')
        role_id=get_role_id(cur,'设计人员')
        for (user_name,login_name) in lst:
            if user_duplication_check(cur,login_name,user_name):
                user_id=add_user(cur,login_name,user_name)
                add_user_group(cur,user_id,group_id)
                add_user_role(cur,user_id,role_id)
                print(login_name,user_name,user_id,group_id,role_id)
        conn.commit()


if __name__=="__main__":
    run()