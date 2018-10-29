# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import os,pymssql,re

def main():
    path=r'D:\桌面\平台\新建文件夹\工作手册'
    file1=r'D:\桌面\平台\新建文件夹\工作簿1.xlsx'
    filelist=os.listdir(path)#该文件夹下所有的文件（包括文件夹）
    with pymssql.connect(host="10.1.42.102", user="qifangchen", password="qfc23834358",database="test2") as conn:
        cur=conn.cursor()
        cur.execute('drop table  Table_1')
        cur.execute('CREATE TABLE [dbo].[Table_1]([id] [smallint] IDENTITY(1,1) NOT NULL,[stage] [nvarchar](50) NULL,[modname] [nvarchar](50) NULL,'
                    '[lv1] [nvarchar](100) NULL,[lv2] [nvarchar](100) NULL,[lv3] [nvarchar](100) NULL,[content] [nvarchar](500) NULL,'
                    '[tongyong] [smallint] NULL,[zhuzhai] [smallint] NULL,[gongjian] [smallint] NULL,[zongheti] [smallint] NULL,[dist] [nvarchar](1000) NULL) ON [PRIMARY]')
        for file in filelist:
            file=os.path.join(path,file)
            parse1(cur,file)
        parse2(cur,file1)
        conn.commit()

def merge(worksheet):
    #查询合并的单元格
    m_list = worksheet.merged_cells
    #判断单元格生成坐标输出到list
    merge_all_list = []
    for m_area in m_list:
        # 合并单元格的起始行坐标、终止行坐标
        r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
        if (r1 != r2 and c1 != c2):
            row_col = [(x, y) for x in range(r1, r2 + 1) for y in range(c1, c2 + 1)]
            merge_all_list.append(row_col)
        elif (r1 == r2 and c1 != c2):  # or (r1 != r2 and c1 == c2):
            col = [(r1, n) for n in range(c1, c2 + 1)]
            merge_all_list.append(col)
        elif (r1 != r2 and c1 == c2):
            row = [(m, c1) for m in range(r1, r2 + 1)]
            merge_all_list.append(row)
    return merge_all_list

def get_head(x,y,merges):
    for merge in merges:
        if (x,y) in merge:
            return merge[0]
    return (x,y)

def get_value(row,col,merges,worksheet):
    (row,col)=get_head(row,col,merges)
    return worksheet.cell(row, col).value

def purge(str):
    if str:
        return str.replace('\n',' ').replace(' ','')
    else:
        return str 

def sheet2stage(sheet):
    if sheet=='方案3690配合0~30%':
        return '0~30阶段'
    elif sheet=='方案3690配合30%~60%':
        return '30~60阶段'
    elif sheet=='方案3690配合60%~90%':
        return '60~90阶段'
    elif sheet=='方案3690配合90%~100%':
        return '90~100阶段'
    else:
        print(sheet+'未获得阶段')
        return '未获得阶段'
def get_type(tempfilename):
    if tempfilename[0:2]=='住宅':
        return (0,1,0,0)
    elif tempfilename[0:2]=='公建':
        return (0,0,1,0)
    else:
        return (0,0,0,0)

def get_type2(value):
    if value=='' or value==0 or value is None:
        return 0
    else:
        return 1
    
def get_lines(merges):
    for merge in merges:
        if (5,2) in merge:
            return merge
'''     
def insert_line2(cur,stage,modname,lv1,lv2,lv3,content,tongyong,zhuzhai,gongjian,zongheti,dist):
    try:
        cur.execute(r'SELECT id,zhuzhai,gongjian,zongheti FROM Table_1 WHERE dist=%s',(dist))
        res=cur.fetchone()
        if res:
            (id1,zhuzhai1,gongjian1,zongheti1)=res
            zhuzhai+=zhuzhai1
            gongjian+=gongjian1
            zongheti+=zongheti1
            cur.execute(r'update Table_1 set zhuzhai=%s,gongjian=%s,zongheti=%s where id=%s',(zhuzhai,gongjian,zongheti,id1))
        else:
            cur.execute('insert into Table_1(stage,modname,lv1,lv2,lv3,[content],tongyong,zhuzhai,gongjian,zongheti,dist) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                    (stage,modname,lv1,lv2,lv3,content,tongyong,zhuzhai,gongjian,zongheti,dist))
    except pymssql.OperationalError:
        print(stage,modname,lv1,lv2,lv3,content,tongyong,zhuzhai,gongjian,zongheti)
'''

def insert_line(cur,stage,modname,lv1,lv2,lv3,content,tongyong,zhuzhai,gongjian,zongheti,dist):
    try:
        cur.execute('insert into Table_1(stage,modname,lv1,lv2,lv3,[content],tongyong,zhuzhai,gongjian,zongheti,dist) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                    (stage,modname,lv1,lv2,lv3,content,tongyong,zhuzhai,gongjian,zongheti,dist))
    except pymssql.OperationalError:
        print(stage,modname,lv1,lv2,lv3,content,tongyong,zhuzhai,gongjian,zongheti)


def parse1(cur,file):
    (filepath,tempfilename) = os.path.split(file) 
    (modname,extension) = os.path.splitext(tempfilename[3:-1])#modname
    workbook = load_workbook(file)
    for sheet in workbook.sheetnames: #sheet at stage
        stage=sheet2stage(sheet)
        worksheet = workbook[sheet]
        merges=merge(worksheet)
        for (row,col) in get_lines(merges):
            if modname[0:2] in ['方案','建筑']:
                lv1=purge(get_value(row,col+3,merges,worksheet))
                lv2=purge(get_value(row,col+4,merges,worksheet))
                lv3=purge(get_value(row,col+5,merges,worksheet))
                content=purge(get_value(row,col+6,merges,worksheet))
            else:
                lv1=purge(get_value(row,col+4,merges,worksheet))
                lv2=purge(get_value(row,col+5,merges,worksheet))
                lv3=purge(get_value(row,col+7,merges,worksheet))
                content=purge(get_value(row,col+8,merges,worksheet))
            if lv1 is None:
                continue
            if lv3==content or lv3 is None:
                lv3=''
            if lv2==lv1 or lv2 is None:
                lv2=lv1
            (tongyong,zhuzhai,gongjian,zongheti)=get_type(tempfilename)
            if content is None:
                cur.execute(r'select count(*) from Table_1 where stage=%s and modname=%s and lv1=%s and lv2=%s and lv3=%s and tongyong=%s and zhuzhai=%s and gongjian=%s and zongheti=%s',
                            (stage,modname,lv1,lv2,lv3,tongyong,zhuzhai,gongjian,zongheti))
                if cur.fetchone()[0]>0:
                    print('content is none')
                    print((stage,modname,lv1,lv2,lv3,tongyong,zhuzhai,gongjian,zongheti))
                    continue
                else:
                    content=lv2
            dist=''.join([modname,lv1,lv2,lv3,content])
            dist=''.join(re.findall(u'[\u4e00-\u9fff]+', dist))
            dist=''.join([stage,dist])
            insert_line(cur,stage,modname,lv1,lv2,lv3,content,tongyong,zhuzhai,gongjian,zongheti,dist)


def parse2(cur,file):
    workbook = load_workbook(file)
    worksheet=workbook['Sheet1']
    for row in range(2,129):
        stage=worksheet.cell(row, 1).value
        modname=worksheet.cell(row, 2).value
        lv1=purge(worksheet.cell(row, 3).value)
        lv2=purge(worksheet.cell(row, 4).value)
        lv3=''
        content=purge(worksheet.cell(row, 5).value)
        tongyong=get_type2(worksheet.cell(row, 6).value)
        zhuzhai=get_type2(worksheet.cell(row, 7).value)
        gongjian=get_type2(worksheet.cell(row, 8).value)
        zongheti=get_type2(worksheet.cell(row, 9).value)
        dist=''.join([modname,lv1,lv2,lv3,content])
        dist=''.join(re.findall(u'[\u4e00-\u9fff]+', dist))
        dist=''.join([stage,dist])
        insert_line(cur,stage,modname,lv1,lv2,lv3,content,tongyong,zhuzhai,gongjian,zongheti,dist)


if __name__=='__main__':
    main()
        
